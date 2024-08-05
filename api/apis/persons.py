from fastapi import APIRouter, HTTPException, UploadFile, status
from fastapi.responses import JSONResponse
from typing import Optional, List, Union
from pydantic import BaseModel, ValidationError, validator
from uuid import UUID
from datetime import datetime, date
import logging
import io
import openpyxl
from dao import persons as players, teams
from .teams import TeamBase, SeasonTeamOut2

router = APIRouter()

class PersonBase(BaseModel):
    first_name: str
    last_name: str
    person_type: Optional[int]
    team_id: Optional[UUID]

class Height(BaseModel):
    feet: Optional[int]
    inches: Optional[int]
    
class PublicPlayerOut(PersonBase):
    id: UUID
    height: Optional[Height]
    player_number: Optional[int]
    position: Optional[str]
    age: Optional[int]
    # season_roster: Optional[List[SeasonTeamOut2]]

class PlayerOut(PersonBase):
    id: UUID
    age: Optional[int]
    height: Optional[Height]
    player_number: Optional[int]
    position: Optional[str]
    age: Optional[int]
    season_roster: Optional[List[SeasonTeamOut2]]

class PlayerIn(PersonBase):
    id: Optional[UUID]
    season_roster: List[SeasonTeamOut2]
    age: Union[int, str]
    height: Optional[Height]
    person_type: Union[int, str]
    player_number: Optional[int]
    position: Optional[str]

class ImportPlayer(BaseModel):
    first_name: str
    last_name: str
    team_slug: str
    year: str
    level_name: str
    age: int
    height: Optional[Height]
    player_number: Optional[int]
    position: Optional[str]
    grade: int


#TODO: Move to Rosters
@router.get('/getPlayers/{slug}', response_model=List[PublicPlayerOut], summary='Get a teams players', tags=['players'])
def get_team_players(slug):
    return players.get_team_list(slug)

@router.get('/getPlayers', summary="Get all players", tags=['players']  )
async def get_all_players():
    rosters = await players.get_list(person_type='Player')
    # print(rosters)
    return rosters


@router.post('/addPlayer', tags=['players'])
def add_player(player: PlayerIn):
    print(player)
    players.create_player(player)
    return {200: "Success"}

@router.post('/updatePlayer/{id}', summary="Update a player", tags=['players', 'rosters'])
@router.put('/updatePlayer/{id}', summary="Update a player", tags=['players', 'rosters'])
def update_player(id, player: PlayerIn):
    try:
        players.update(id, player)
    except Exception as exc:
        print(str(exc))
        return HTTPException(status_code = 400, detail= "Error Message")
    return {200: "Success"}

@router.post('/addPlayerToRoster', tags=['players'])
def add_to_roster():
    pass

@router.get('/getRoster/{season_team}', summary="Get a leveled team roster", tags=['players', 'rosters'])
def get_team_roster(season_team: UUID):
    team = teams._get_slug_by_level_id(season_team).get('slug')
    return players.get_team_list(team, season_team) 
    # try:
    #     roster = players.get_team_list(team, season_team)
    # except HTTPException:
    #     raise 
    # return roster


#TODO: Move to Rosters
@router.get('/getAdminPlayers/{slug}', summary='Get a teams players', tags=['players'])
def get_team_players(slug):
    try:
        player_list = players.get_team_list(slug)
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return player_list


@router.get('/getAdminPlayers', summary="Get all players", tags=['players']  )
def get_all_players():
    try:
        player_list = players.get_list(person_type='Player')
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return player_list


@router.post('/teamFile/{team_slug}/{year}')
async def create_team_file(file: UploadFile, team_slug: str, year: str):
    error = False
    #TODO: Needs season_roster: List[SeasonTeamOut2]
    if file.filename.endswith('.xlsx'):
        f = await file.read()
        xlsx = io.BytesIO(f)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        level_name = ws.cell(row=1, column=3).value
        import re
        pattern = re.compile(r"""(\d+)(?:'|’)(?: *(\d+))?""")
        headers = ['player_number', 'age', 'grade', 'position', 'height', 'first_name', 'last_name']
        response_list = []
        # From the excel file players start in A4
        for row in ws.iter_rows(min_row=4):
            row_values = [cell.value for cell in row]
            if set(row_values) != set([None]):
                person = dict(zip(headers, [cell.value for cell in row]))
                feet, inches = re.match(pattern, person['height']).groups()
            
                person['height'] = {
                    'feet': int(feet or 0),
                    'inches': int(inches or 0)
                }
                
                person['level_name'] = level_name.lower()
                person['team_slug'] = team_slug
                person['year'] = year
                person['person_type'] = '1'
                response = players.import_player(ImportPlayer(**person))
                print(response)
                if response['status_code'] != 200:
                    error = True
                response_list.append(response['detail'])
        
        status_code = status.HTTP_200_OK
        if error:
            status_code = status.HTTP_409_CONFLICT

        return JSONResponse(status_code=status_code, content=response_list)